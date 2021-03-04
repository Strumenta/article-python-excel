import unittest
from typing import Tuple, Union

import formulas
from openpyxl import load_workbook
from openpyxl.cell import Cell


class TestFormulas(unittest.TestCase):
    def test_data_only(self):
        wb = load_workbook('data/test_formulas.xlsx', data_only=True)
        try:
            self.assertEqual(wb.active['A1'].value, 42)
            self.assertEqual(wb.active['B1'].value, 21)
            self.assertEqual(wb.active['C1'].value, 31.5)
        finally:
            wb.close()

    def test_calculate(self):
        xl_model = formulas.ExcelModel().loads('data/test_formulas.xlsx').finish()
        solution = xl_model.calculate()
        c1 = solution['\'[TEST_FORMULAS.XLSX]SHEET1\'!C1']
        self.assertEqual(c1.value[0], 31.5)

    def test_formula_function(self):
        func = formulas.Parser().ast("=A1+B1")[1].compile()
        self.assertEqual(3, func(1, 2))

    def test_compute_formula(self):
        wb = load_workbook('data/test_formulas.xlsx')
        self.assertEqual(compute_cell_value(wb.active['C1']), 31.5)
        wb.active['A1'].value = 100
        self.assertEqual(compute_cell_value(wb.active['C1']), 75)
        wb.close()

    def test_compute_formula_range_single_row(self):
        wb = load_workbook('data/test_formulas.xlsx')
        self.assertEqual(compute_cell_value(wb.active['D1']), 94.5)
        wb.active['A1'].value = 100
        self.assertEqual(compute_cell_value(wb.active['D1']), 225)
        wb.close()

    def test_compute_formula_range_multi_row(self):
        wb = load_workbook('data/test_formulas.xlsx')
        self.assertEqual(compute_cell_value(wb.active['D2']), 157.5)
        wb.active['A1'].value = 100
        self.assertEqual(compute_cell_value(wb.active['D2']), 375)
        wb.close()


def has_formula(cell: Cell):
    return isinstance(cell.value, str) and cell.value.startswith('=')


def compute_cell_value(input: Union[Cell, Tuple]):
    if isinstance(input, Tuple):
        return tuple(map(compute_cell_value, input))
    if not has_formula(input):
        return input.value
    func = formulas.Parser().ast(input.value)[1].compile()
    args = []
    sheet = input.parent
    for key in func.inputs.keys():
        args.append(compute_cell_value(sheet[key]))
    return func(*args)
