import unittest
from openpyxl import Workbook


class TestAddRemoveSheets(unittest.TestCase):
    def testEmptySheetNames(self):
        wb = Workbook()
        self.assertEqual(['Sheet'], wb.sheetnames)
        wb.close()

    def testAddSheet(self):
        wb = Workbook()
        wb.create_sheet('Another Sheet', 0)
        self.assertEqual(['Another Sheet', 'Sheet'], wb.sheetnames)
        wb.close()

    def testRemoveSheet(self):
        wb = Workbook()
        del wb['Sheet']
        self.assertEqual([], wb.sheetnames)
        wb.close()


class TestAddRemoveCells(unittest.TestCase):
    def testSetCellAddsRowCol(self):
        wb = Workbook()
        self.assertEqual(wb.active.max_row, 1)
        self.assertEqual(wb.active.max_column, 1)
        wb.active['C3'].value = 12
        self.assertEqual(wb.active.max_row, 3)
        self.assertEqual(wb.active.max_column, 3)
        wb.close()

    def testAddRemoveRows(self):
        wb = Workbook()
        self.assertEqual(wb.active.max_row, 1)
        wb.active['A1'].value = 11
        wb.active.insert_rows(0, 3)
        self.assertEqual(wb.active.max_row, 4)
        self.assertEqual(wb.active['A4'].value, 11)
        wb.active.delete_rows(2, 1)
        self.assertEqual(wb.active.max_row, 3)
        self.assertEqual(wb.active['A3'].value, 11)
        del wb.active['A3']
        self.assertEqual(wb.active.max_row, 2)
        wb.close()

    def testAddRemoveCols(self):
        wb = Workbook()
        self.assertEqual(wb.active.max_column, 1)
        wb.active['A1'].value = 11
        wb.active.insert_cols(0, 3)
        self.assertEqual(wb.active.max_column, 4)
        self.assertEqual(wb.active['D1'].value, 11)
        wb.active.delete_cols(2, 1)
        self.assertEqual(wb.active.max_column, 3)
        self.assertEqual(wb.active['C1'].value, 11)
        del wb.active['C1']
        self.assertEqual(wb.active.max_column, 2)
        wb.close()
